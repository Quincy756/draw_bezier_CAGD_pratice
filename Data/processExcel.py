import matplotlib.pyplot as plt
from PyQt5.QtWidgets import  *
import openpyxl
import os

class ExcelHandle:
    def __init__(self, data=None):
        self.wb = None
        self.currentSheet = None
        self.data = data

    def getData(self):
        return self.data

    def setData(self, data):
        self.data = data

    def getWorkbook(self, file_path, sheetName="Sheet1"):
        try:
            is_fileExist = os.path.exists(file_path)
            if is_fileExist:
                wb = openpyxl.load_workbook(file_path)
            else:
                wb = openpyxl.Workbook()
            self.wb = wb
            # 打开表格
            # 打开表单名，如果表单不存在则创建新表单
            # print("======")
            if sheetName not in self.wb.sheetnames:
                self.currentSheet = self.wb.create_sheet(sheetName)

            self.currentSheet = self.wb[sheetName]
            self.currentSheet = self.wb.active
            # print("-------")
        except Exception as ex:
            self.data = []


    def saveData(self, file_path, sheetName="Sheet1"):

        self.getWorkbook(file_path, sheetName)
        self.currentSheet.cell(1, 1).value = "x"
        self.currentSheet.cell(1, 2).value = "y"
        columns = len(self.data)
        rows = len(self.data[0]) + 1
        print("data")
        for row in range(2, rows+1):
            for col in range(1, columns+1):
                self.currentSheet.cell(row, col).value = self.data[col-1][row-2]
        self.wb.save(file_path)

    # path 表示文件得相对路径
    # 将excel中的数据导入到gui显示的表格中
    def exportData(self, file_path, sheetName="Sheet1"):

        self.getWorkbook(file_path, sheetName)
        rows = self.currentSheet.max_row
        columns = self.currentSheet.max_column
        x_header = self.currentSheet.cell(1, 1).value
        y_header = self.currentSheet.cell(1, 2).value
        print(rows, columns)
        data = []
        print(data)
        for col in range(1, columns+1):
            temp = []
            for row in range(2, rows + 1):
                print(self.currentSheet.cell(row, col).value)
                temp.append(self.currentSheet.cell(row, col).value)
            data.append(temp)
        return data


class processData:
    def __init__(self, data=None):
        self.data = data
        # self.data = [[1, 2, 3], [4, 6, 8]]
        self.excel = ExcelHandle(self.data)

    def saveData(self):
        # print("yes")
        # self.saveDataPath = "..//data//My Data/新建文件.xlsx"
        self.saveDataPath = "Data/"
        filedialog = QFileDialog()
        filedialog.setFileMode(QFileDialog.Directory)
        file_path, file_name = QFileDialog.getSaveFileName(filedialog, "保存", self.saveDataPath,\
                             "Excel Files (*.xlsx);;Figure Files(*.fig);;PNG Files(*.png);;JPEG(*.jpg)")
        if file_path:
            if file_name == "Excel Files (*.xlsx)":
                # 调用main.py 中的获取表格数据点函数
                self.updateFunc[1]()
                print("---------------")
                print(self.data)
                self.excel.saveData(file_path)

            elif file_name == "Figure Files(*.fig)":
                print("yes")

            elif file_name == "JPEG(*.jpg)" or "PNG Files(*.png)":
                self.saveFig(file_path)

    def setFig(self, figure):
        self.figure = figure

    def saveFig(self, file_path):
        plt.savefig(file_path)
        plt.close()


    def setData(self, data):
        self.data = data
        self.excel.setData(data)

    def getData(self):
        return self.data

    def getUpdateFunc(self, func):
        self.updateFunc = func

    def exportData(self):
        # self.importDataPath = "..//data//我的数据/"
        self.exportPath = "Data/"
        filedialog = QFileDialog()
        filedialog.setFileMode(QFileDialog.ExistingFiles)
        file_path, file_name = QFileDialog.getOpenFileName(filedialog, "导入", self.exportPath, \
                                        "Excel Files (*.xlsx);;Figure Files(*.fig)")

        if file_path:
            if file_name == "Excel Files (*.xlsx)":
                self.data = self.excel.exportData(file_path, "Sheet")
                self.updateFunc[0](self.data)

            elif file_name == "Figure Files(*.fig)":
                print("yes")




